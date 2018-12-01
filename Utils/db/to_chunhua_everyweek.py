#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl.reader.excel import load_workbook
import xlsxwriter
import MySQLdb
import datetime
from decimal import Decimal
import os,time

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import traceback



import sys
reload(sys)
sys.setdefaultencoding('utf8')

def encode_data(data):
    r = []
    for row in data:
        l = []
        for s in row:
            if isinstance(s, unicode):
                pass
            elif isinstance(s, datetime.datetime):
                s = s.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(s, datetime.date):
                s = s.strftime("%Y-%m-%d")
            elif isinstance(s, datetime.time):
                s = str(s)
            elif isinstance(s, datetime.timedelta):
                s = str(s)
            elif isinstance(s, long):
                s=str(s)
            elif isinstance(s, float):
                s=str(s)
            else :
                pass
            l.append(s)
        r.append(l)
    return r


def write_excel(db,sql,file_path,title):

    ## 查询数据
    db = db
    cursor = db.cursor()
    sql = sql
    try:
        cursor.execute(sql)
        data = cursor.fetchall()
        #db.close()
    except:
        print("Error: unable to fecth data")
        db.close()
        return


    data = encode_data(data)



    #workbook.close() 不关闭是否能持续写入到一个excel中


def sendmail(subject,msg,toaddrs,fromaddr,smtpaddr,password,fujian):
    print("sendmail tooo")


def get_date():
    """
    获取上个月第一天的日期 和 本月第一天的日期
    :return: 返回上述日期的数组
    """
    today=datetime.datetime.today()
    year=today.year
    month=today.month
    if month == 1:
        stop_time=str(year) + '-0' + str(month) + '-01'
        month_before = 12
        year_before = year -1
        start_time=str(year_before) + '-' + str(month_before) + '-01'
    elif month > 1 and month < 10 :
        stop_time=str(year) + '-0'  + str(month) + '-01'
        month_before = month -1
        start_time=str(year) + '-0' + str(month_before) + '-01'
    elif month == 10 :
        stop_time=str(year) + '-' + str(month) + '-01'
        month_before = month -1
        start_time=str(year) + '-0' + str(month_before) + '-01'
    else  :
        stop_time=str(year) + '-'  + str(month) + '-01'
        month_before = month -1
        start_time=str(year) + '-'  + str(month_before) + '-01'

    return (start_time,stop_time)



if __name__ == '__main__':


    """ 数据库连接"""
    db = MySQLdb.connect("10.26.169.169", "read_only", "chaboshi876", "db_www", 3308 ,charset="utf8")


    """ 保存成excel ,worksheet都得一次性在这里初始化"""
    # start_time = time.strftime('%Y-%m-%d', time.localtime(time.time()-592200))
    # #print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()-592200)))
    # stop_time = time.strftime('%Y-%m-%d',time.localtime(time.time()))
    #file_path_base = '/opt/soft/python_script/jingpin/'
    start_time = '2018-08-01'
    stop_time ='2018-08-02'
    file_path_base = 'D:/jingpin/'
    file_path1 = file_path_base + stop_time +  "_clew.xlsx"
    workbook = xlsxwriter.Workbook(file_path1)
    worksheet1 = workbook.add_worksheet('上周线索')

    worksheet2 = workbook.add_worksheet('本月线索')
    #worksheet3 = workbook.add_worksheet('本月查博士检测')

    """调整一下列的宽度"""
    worksheet1.set_column('B:AZ',22)

    """保存标题"""
    row = 0
    col = 0
    title = ['id','58单号','添加时间','电话','车架号','类型','deleted','线索备注','CBS订单号','下单时间','产品类型' ,'产品子类型','订单主状态','订单子状态','用户id','下单账号','用户昵称','用户类型' ,'联系人','联系地址','联系人电话','取消原因','备注' ]
    worksheet1.write_row(row,col,title)
    """ 拼写SQL语句 """
    sql_1="""
        SELECT
                t.id,
                t.id_detection,
                t.add_time,
                t.phone,
                D.vin,
                t.type,
                t.deleted,
                tcr.remark '线索备注',
                D.order_no '订单号' ,
                D.add_time '下单时间',
                case when B.order_type=5 THEN '事故保'
			          when B.order_type=6 THEN '车况保'
                      when B.order_type=7 THEN '车价保' END  '产品类型' ,
                case  when  B.order_type='5' AND  D.type='1'  THEN '本地'
                      when  B.order_type='5' AND  D.type='2'  THEN '异地'
                      when  B.order_type='5' AND  D.type='3'  THEN '本地初检'
                      when  B.order_type='5' AND  D.type='4'  THEN '本地复检'
                      when  B.order_type='5' AND  D.type='5'  THEN '异地初检'
                      when  B.order_type='5' AND  D.type='6'  THEN '异地复检'
                      when  B.order_type='5' AND  D.type='7'  THEN '58放心车'
                      when  B.order_type='5' AND  D.type='8'  THEN '金融检测'
                      when  B.order_type='5' AND  D.type='9'  THEN '58_A套餐'
                      when  B.order_type='5' AND  D.type='10'  THEN '58异地检'
                      when  B.order_type='5' AND  D.type='11'  THEN '远程检测'
                      when  B.order_type='5' AND  D.type='12'  THEN '延保前置检测'
                      when  B.order_type='6' AND  D.type='1'  THEN 'A套餐'
                      when  B.order_type='6' AND  D.type='2'  THEN 'B套餐'
                      when  B.order_type='6' AND  D.type='3'  THEN 'A套餐初检'
                      when  B.order_type='6' AND  D.type='4'  THEN 'A套餐复检'
                      when  B.order_type='6' AND  D.type='5'  THEN 'B套餐初检'
                      when  B.order_type='6' AND  D.type='6'  THEN 'B套餐复检'
                      when  B.order_type='6' AND  D.type='7'  THEN '58延保'
                      when  B.order_type='6' AND  D.type='8'  THEN '58_B套餐'
                      when  B.order_type='6' AND  D.type='9'  THEN '延保检测'  END '产品子类型',
                case when B.order_status=1000100 THEN  '待确认'
                    when B.order_status=1000200 THEN  '待付款'
                    when B.order_status=1000300 THEN  '已付款'
                    when B.order_status=1000400 THEN  '已完成'
                    when B.order_status=1000500 THEN  '已取消'
                    when B.order_status=1000600 THEN  '已退款' END  '订单主状态',
                case when D.`status`=10010 THEN   '待分配'
                    when D.`status`=10020 THEN '已分配'
                    when D.`status`=10030 THEN '待付款'
                    when D.`status`=10040 THEN '待检测'
                    when D.`status`=10050 THEN '待承保'
                    when D.`status`=10055 THEN '审核中'
                    when D.`status`=10060 THEN '审核通过'
                    when D.`status`=10070 THEN '审核不通过'
                    when D.`status`=10080 THEN '第三方承保失败（保留状态）'
                    when D.`status`=10090 THEN '已取消' END '订单子状态',
				B.user_id '用户id',
                C.mobile '下单账号',
                C.nick_name,
                CASE WHEN C.user_type =0 THEN '个人'
                     WHEN C.user_type =1 THEN '企业'
                     WHEN C.user_type =2 THEN '接口'
                     WHEN C.user_type =3 THEN '大客户'
                     WHEN C.user_type =4 THEN '经销商'
                     WHEN C.user_type =5 THEN '测试'
                     WHEN C.user_type =6 THEN '销售'
                     WHEN C.user_type =7 THEN '4s店'
                     WHEN C.user_type =8 THEN '中介' END  '用户类型' ,
                D.name '联系人',
                D.address '联系地址',
                D.mobile '联系人电话',
                E.message '取消原因',
                replace(SUBSTRING_INDEX(D.remark,'content":"',-1),'"}]','') '备注'


            FROM  t_detection_clew t
            LEFT JOIN t_detection_clew_record tcr  ON t.id=tcr.clew_id
            LEFT JOIN t_insurence_detail D ON t.id_detection =D.thirdpart_orderid
            LEFT JOIN t_insurence_detail_action E ON D.order_no=E.order_no AND E.`status`=10090
            LEFT JOIN t_order B ON D.order_no=B.order_no
            LEFT JOIN t_user C ON B.user_id=C.id
            WHERE  t.add_time   BETWEEN '""" + start_time + """' and '"""  + stop_time  + """' ; """

    print(sql_1)
    cursor = db.cursor()
    ''' 抓取数据 '''
    try:
        cursor.execute(sql_1)
        data = cursor.fetchall()
        #db.close()
    except:
        print("Error: unable to fecth data")
        db.close()

    row = 1
    col = 0

    data = encode_data(data)
    for line in data:
        worksheet1.write_row(row,col,line)
        row += 1
    cursor.close()






    db.close()
    workbook.close()



    ''' 将获取的结果，发邮件给同事 '''
    fromaddr = "system_call@chaboshi.cn"
    smtpaddr = "smtp.exmail.qq.com"
    password = "QXPtfEyuJm6uunTd"
    msg = "请查收附件。"

    """ 发给红霞姐的"""
    #toaddrs1 = ["weiyangbo@chaboshi.cn"]
    toaddrs1 = ["wangfeiyue@chaboshi.cn"]
    subject1 = "统计周报-运营"
    fujian=file_path1
    #sendmail(subject1,msg,toaddrs1,fromaddr,smtpaddr,password,fujian)





#ws.max_row + 1   --- 代表最后一行的行数值

#   AND A.user_id  NOT in ( 118,38888,45655)